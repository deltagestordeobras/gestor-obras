from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors, pagesizes
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from database.connection import BASE_DIR
from services.insumos import carregar_insumos
from services.lancamento import carregar_lancamentos
from services.status import STATUS_PAGO, normalizar_status


COLUNAS_RELATORIO = [
    "Descrição", "Data da nota", "Data de pagamento", "Categoria", "Etapa",
    "Fornecedor", "Quantidade", "Valor unitário", "Valor total", "Status",
    "Nº nota", "Obra", "Comprovante", "NotaID",
]


def _fornecedor(nota):
    return str(nota.get("Fornecedor") or nota.get("Descrição") or nota.get("Descrição") or "").strip()


def _valor_nota(nota):
    valor = pd.to_numeric(nota.get("Valor"), errors="coerce")
    return abs(float(valor)) if pd.notna(valor) else 0.0


def _linha_base(nota, obra):
    status = normalizar_status(nota.get("Status"))
    return {
        "Data da nota": pd.to_datetime(nota.get("Entrada Nota"), errors="coerce"),
        "Data de pagamento": pd.to_datetime(nota.get("Data Pagto"), errors="coerce"),
        "Categoria": nota.get("Categoria") or "Sem categoria",
        "Etapa": nota.get("Etapa") or "Sem etapa",
        "Fornecedor": _fornecedor(nota) or "Não informado",
        "Status": "Pago" if status == STATUS_PAGO else "A pagar",
        "Nº nota": nota.get("Nº Nota") or nota.get("NÂº Nota") or "",
        "Obra": obra,
        "Comprovante": nota.get("Foto") or "",
        "NotaID": nota.get("ID") or "",
    }


def consolidar_despesas(obra):
    notas = carregar_lancamentos(obra)
    insumos = carregar_insumos(obra)
    if notas is None or notas.empty:
        return pd.DataFrame(columns=COLUNAS_RELATORIO)

    notas = notas.copy()
    notas["Valor"] = pd.to_numeric(notas["Valor"], errors="coerce").fillna(0)
    notas = notas[notas["Valor"] < 0]
    linhas = []

    for _, nota in notas.iterrows():
        base = _linha_base(nota, obra)
        itens = insumos[insumos["NotaID"].astype(str) == str(nota.get("ID"))].copy() if not insumos.empty else pd.DataFrame()
        valor_nota = _valor_nota(nota)

        if itens.empty:
            linhas.append({
                **base,
                "Descrição": _fornecedor(nota) or str(nota.get("Categoria") or "Despesa"),
                "Quantidade": 1.0,
                "Valor unitário": valor_nota,
                "Valor total": valor_nota,
            })
            continue

        itens["Quantidade"] = pd.to_numeric(itens["Quantidade"], errors="coerce").fillna(0)
        itens["ValorUnitario"] = pd.to_numeric(itens["ValorUnitario"], errors="coerce").fillna(0)
        itens["ValorTotal"] = pd.to_numeric(itens["ValorTotal"], errors="coerce").fillna(0)
        for _, item in itens.iterrows():
            linhas.append({
                **base,
                "Descrição": item.get("Material") or "Item sem descrição",
                "Etapa": item.get("Etapa") or base["Etapa"],
                "Quantidade": float(item["Quantidade"]),
                "Valor unitário": float(item["ValorUnitario"]),
                "Valor total": float(item["ValorTotal"]),
            })

        diferenca = round(valor_nota - float(itens["ValorTotal"].sum()), 2)
        if abs(diferenca) >= 0.01:
            linhas.append({
                **base,
                "Descrição": "Complemento/Ajuste da nota",
                "Quantidade": 1.0,
                "Valor unitário": diferenca,
                "Valor total": diferenca,
            })

    df = pd.DataFrame(linhas, columns=COLUNAS_RELATORIO)
    if not df.empty:
        df = df.sort_values("Data da nota", ascending=False, na_position="last").reset_index(drop=True)
    return df


def aplicar_filtros(df, inicio=None, fim=None, categoria=None, etapa=None, fornecedor=None, status=None):
    resultado = df.copy()
    if inicio:
        resultado = resultado[resultado["Data da nota"].dt.date >= pd.Timestamp(inicio).date()]
    if fim:
        resultado = resultado[resultado["Data da nota"].dt.date <= pd.Timestamp(fim).date()]
    for coluna, valor in (
        ("Categoria", categoria), ("Etapa", etapa), ("Fornecedor", fornecedor), ("Status", status)
    ):
        if valor and valor != "Todos":
            resultado = resultado[resultado[coluna] == valor]
    return resultado


def caminho_comprovante(caminho):
    if not caminho or caminho == "-":
        return None
    candidato = Path(str(caminho))
    if not candidato.is_absolute():
        candidato = BASE_DIR / candidato
    candidato = candidato.resolve()
    raiz = BASE_DIR.resolve()
    if raiz not in candidato.parents:
        return None
    return candidato if candidato.exists() and candidato.is_file() else None


def gerar_excel_despesas(df, obra):
    buffer = BytesIO()
    exportacao = df.drop(columns=["Comprovante", "NotaID"], errors="ignore").copy()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        exportacao.to_excel(writer, index=False, sheet_name="Despesas")
        resumo = pd.DataFrame({
            "Indicador": ["Obra", "Total geral", "Quantidade de itens"],
            "Valor": [obra, float(df["Valor total"].sum()), len(df)],
        })
        resumo.to_excel(writer, index=False, sheet_name="Resumo")
        ws = writer.book["Despesas"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for coluna in ws.columns:
            largura = min(38, max(12, max(len(str(c.value or "")) for c in coluna) + 2))
            ws.column_dimensions[coluna[0].column_letter].width = largura
    return buffer.getvalue()


def gerar_pdf_despesas(df, obra):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=pagesizes.landscape(pagesizes.A4),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"<b>Materiais e Despesas da Obra</b> - {obra}", styles["Title"]),
        Spacer(1, 10),
        Paragraph(f"Total geral: R$ {df['Valor total'].sum():,.2f} | Itens: {len(df)}", styles["Normal"]),
        Spacer(1, 10),
    ]
    cabecalho = ["Descrição", "Data", "Categoria", "Etapa", "Fornecedor", "Qtd.", "Unitário", "Total", "Status", "Nota"]
    dados = [cabecalho]
    for _, row in df.iterrows():
        dados.append([
            str(row["Descrição"])[:34], row["Data da nota"].strftime("%d/%m/%Y") if pd.notna(row["Data da nota"]) else "",
            str(row["Categoria"])[:18], str(row["Etapa"])[:18], str(row["Fornecedor"])[:24],
            f"{row['Quantidade']:,.2f}", f"R$ {row['Valor unitário']:,.2f}",
            f"R$ {row['Valor total']:,.2f}", str(row["Status"]), str(row["Nº nota"])[:14],
        ])
    tabela = Table(dados, repeatRows=1, colWidths=[125, 55, 80, 75, 100, 45, 65, 65, 55, 60])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    return buffer.getvalue()
